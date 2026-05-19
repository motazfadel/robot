import sqlite3
import subprocess
import shutil
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
import json
import os
from pathlib import Path
import re
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from urllib import error, request


def _get_app_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def _load_simple_env_file(env_path: Path) -> None:
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        if not key or key in os.environ:
            continue
        os.environ[key] = value.strip().strip('"').strip("'")


def _load_env_config() -> None:
    base_dir = _get_app_base_dir()
    candidates = (
        base_dir / ".env",
        base_dir.parent / ".env",
        Path.cwd() / ".env",
        base_dir / ".env.example",
        base_dir.parent / ".env.example",
        Path.cwd() / ".env.example",
    )
    seen: set[Path] = set()
    for candidate in candidates:
        resolved = candidate.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        _load_simple_env_file(resolved)


_load_env_config()


def _load_voice_input_modules():
    try:
        import numpy as np  # type: ignore
        import sounddevice as sd  # type: ignore
        import speech_recognition as sr  # type: ignore
    except Exception:
        return None, None, None
    return sr, np, sd


def _load_workbook_class():
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception:
        return None
    return Workbook


DB_PATH = _get_app_base_dir() / "joe_desktop.db"


@dataclass
class DebtSummary:
    due_today: int
    overdue: int
    total_open_amount: float


@dataclass
class ChatMessage:
    sender: str
    content: str


@dataclass
class CommandResult:
    reply: str
    refresh_ui: bool = False


@dataclass
class ParsedCommand:
    intent: str
    person_name: str = ""
    vendor_name: str = ""
    title: str = ""
    item_name: str = ""
    amount: float = 0.0
    currency: str = "USD"
    due_date: str = ""
    bill_date: str = ""
    category: str = ""
    notes: str = ""


class Database:
    def __init__(self, db_path: Path) -> None:
        self.conn = sqlite3.connect(db_path)
        self.conn.row_factory = sqlite3.Row
        self._create_tables()

    def _create_tables(self) -> None:
        self.conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS reminders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                due_date TEXT NOT NULL,
                notes TEXT
            );

            CREATE TABLE IF NOT EXISTS debts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                person_name TEXT NOT NULL,
                amount REAL NOT NULL,
                currency TEXT NOT NULL,
                due_date TEXT NOT NULL,
                notes TEXT,
                is_paid INTEGER NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS bills (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vendor_name TEXT NOT NULL,
                amount REAL NOT NULL,
                currency TEXT NOT NULL,
                bill_date TEXT NOT NULL,
                category TEXT NOT NULL,
                is_paid INTEGER NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS shopping_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                item_name TEXT NOT NULL,
                added_by TEXT,
                is_done INTEGER NOT NULL DEFAULT 0
            );
            """
        )
        self.conn.commit()

    def add_reminder(self, title: str, due_date: str, notes: str) -> None:
        self.conn.execute(
            "INSERT INTO reminders(title, due_date, notes) VALUES (?, ?, ?)",
            (title, due_date, notes or None),
        )
        self.conn.commit()

    def add_debt(self, person_name: str, amount: float, currency: str, due_date: str, notes: str) -> None:
        self.conn.execute(
            """
            INSERT INTO debts(person_name, amount, currency, due_date, notes)
            VALUES (?, ?, ?, ?, ?)
            """,
            (person_name, amount, currency, due_date, notes or None),
        )
        self.conn.commit()

    def add_bill(self, vendor_name: str, amount: float, currency: str, bill_date: str, category: str) -> None:
        self.conn.execute(
            """
            INSERT INTO bills(vendor_name, amount, currency, bill_date, category)
            VALUES (?, ?, ?, ?, ?)
            """,
            (vendor_name, amount, currency, bill_date, category),
        )
        self.conn.commit()

    def add_shopping_item(self, item_name: str, added_by: str) -> None:
        self.conn.execute(
            "INSERT INTO shopping_items(item_name, added_by) VALUES (?, ?)",
            (item_name, added_by or None),
        )
        self.conn.commit()

    def mark_debt_paid(self, debt_id: int) -> None:
        self.conn.execute(
            "UPDATE debts SET is_paid = 1 WHERE id = ?",
            (debt_id,),
        )
        self.conn.commit()

    def mark_shopping_item_done(self, item_id: int) -> None:
        self.conn.execute(
            "UPDATE shopping_items SET is_done = 1 WHERE id = ?",
            (item_id,),
        )
        self.conn.commit()

    def delete_reminder(self, reminder_id: int) -> None:
        self.conn.execute("DELETE FROM reminders WHERE id = ?", (reminder_id,))
        self.conn.commit()

    def delete_debt(self, debt_id: int) -> None:
        self.conn.execute("DELETE FROM debts WHERE id = ?", (debt_id,))
        self.conn.commit()

    def delete_bill(self, bill_id: int) -> None:
        self.conn.execute("DELETE FROM bills WHERE id = ?", (bill_id,))
        self.conn.commit()

    def delete_shopping_item(self, item_id: int) -> None:
        self.conn.execute("DELETE FROM shopping_items WHERE id = ?", (item_id,))
        self.conn.commit()

    def fetch_reminders(self) -> list[sqlite3.Row]:
        return self.conn.execute(
            "SELECT * FROM reminders ORDER BY due_date ASC, id DESC"
        ).fetchall()

    def fetch_debts(self) -> list[sqlite3.Row]:
        return self.conn.execute(
            "SELECT * FROM debts ORDER BY due_date ASC, id DESC"
        ).fetchall()

    def fetch_bills(self) -> list[sqlite3.Row]:
        return self.conn.execute(
            "SELECT * FROM bills ORDER BY bill_date DESC, id DESC"
        ).fetchall()

    def fetch_all_bills(self) -> list[sqlite3.Row]:
        return self.conn.execute(
            "SELECT * FROM bills ORDER BY bill_date DESC, id DESC"
        ).fetchall()

    def fetch_shopping_items(self) -> list[sqlite3.Row]:
        return self.conn.execute(
            "SELECT * FROM shopping_items WHERE is_done = 0 ORDER BY id DESC"
        ).fetchall()

    def fetch_summary(self) -> DebtSummary:
        today = date.today().isoformat()
        due_today = self.conn.execute(
            "SELECT COUNT(*) FROM debts WHERE is_paid = 0 AND due_date = ?",
            (today,),
        ).fetchone()[0]
        overdue = self.conn.execute(
            "SELECT COUNT(*) FROM debts WHERE is_paid = 0 AND due_date < ?",
            (today,),
        ).fetchone()[0]
        total_open_amount = self.conn.execute(
            "SELECT COALESCE(SUM(amount), 0) FROM debts WHERE is_paid = 0"
        ).fetchone()[0]
        return DebtSummary(due_today=due_today, overdue=overdue, total_open_amount=total_open_amount)

    def fetch_home_snapshot(self) -> dict[str, int]:
        today = date.today().isoformat()
        return {
            "today_reminders": self.conn.execute(
                "SELECT COUNT(*) FROM reminders WHERE due_date = ?",
                (today,),
            ).fetchone()[0],
            "open_bills": self.conn.execute(
                "SELECT COUNT(*) FROM bills WHERE is_paid = 0"
            ).fetchone()[0],
            "shopping_items": self.conn.execute(
                "SELECT COUNT(*) FROM shopping_items WHERE is_done = 0"
            ).fetchone()[0],
        }


class LocalAIParser:
    def __init__(self) -> None:
        self.mode = os.getenv("JOE_AI_MODE", "auto").strip().lower()
        self.endpoint = os.getenv("JOE_OLLAMA_URL", "http://127.0.0.1:11434/api/generate")
        self.model = os.getenv("JOE_OLLAMA_MODEL", "qwen2.5:3b")
        self.openai_api_key = os.getenv("OPENAI_API_KEY", "").strip()
        self.openai_model = os.getenv("JOE_OPENAI_MODEL", "gpt-5.4-mini").strip()
        self.openai_endpoint = os.getenv("JOE_OPENAI_URL", "https://api.openai.com/v1/responses").strip()
        self.last_provider = "startup"
        self.last_error = ""

    def _set_status(self, provider: str, error_message: str = "") -> None:
        self.last_provider = provider
        self.last_error = error_message

    def _should_use_online(self) -> bool:
        if self.mode == "online":
            if not self.openai_api_key:
                self._set_status("local", "OpenAI key missing")
                return False
            return True
        if self.mode == "local":
            self._set_status("local")
            return False
        if self.openai_api_key:
            return True
        self._set_status("local", "OpenAI key missing")
        return False

    def parse(self, text: str) -> ParsedCommand | None:
        if self._should_use_online():
            response_text = self._call_openai(self._build_prompt(text))
            if response_text:
                try:
                    result = json.loads(response_text)
                    return ParsedCommand(
                        intent=result.get("intent", "unknown"),
                        person_name=result.get("person_name", "") or "",
                        vendor_name=result.get("vendor_name", "") or "",
                        title=result.get("title", "") or "",
                        item_name=result.get("item_name", "") or "",
                        amount=float(result.get("amount", 0) or 0),
                        currency=result.get("currency", "USD") or "USD",
                        due_date=result.get("due_date", "") or "",
                        bill_date=result.get("bill_date", "") or "",
                        category=result.get("category", "") or "",
                        notes=result.get("notes", "") or "",
                    )
                except Exception:
                    self._set_status("openai", "OpenAI returned invalid JSON")
            elif self.mode == "online":
                self._set_status("local", self.last_error or "OpenAI request failed")
        else:
            if self.mode == "online":
                self._set_status("local", self.last_error or "OpenAI unavailable")

        payload = {
            "model": self.model,
            "prompt": self._build_prompt(text),
            "stream": False,
            "format": "json",
            "options": {"temperature": 0.1},
        }
        req = request.Request(
            self.endpoint,
            data=json.dumps(payload).encode("utf-8"),
            headers={"Content-Type": "application/json"},
            method="POST",
        )

        try:
            with request.urlopen(req, timeout=25) as response:
                raw = response.read().decode("utf-8")
        except (error.URLError, TimeoutError, OSError) as exc:
            self._set_status("local", f"Ollama request failed: {exc.__class__.__name__}")
            return None

        try:
            envelope = json.loads(raw)
            model_text = envelope.get("response", "{}")
            result = json.loads(model_text)
            self._set_status("ollama")
            return ParsedCommand(
                intent=result.get("intent", "unknown"),
                person_name=result.get("person_name", "") or "",
                vendor_name=result.get("vendor_name", "") or "",
                title=result.get("title", "") or "",
                item_name=result.get("item_name", "") or "",
                amount=float(result.get("amount", 0) or 0),
                currency=result.get("currency", "USD") or "USD",
                due_date=result.get("due_date", "") or "",
                bill_date=result.get("bill_date", "") or "",
                category=result.get("category", "") or "",
                notes=result.get("notes", "") or "",
            )
        except Exception:
            self._set_status("local", "Ollama returned invalid JSON")
            pass

        return None

    def availability_hint(self) -> str:
        return f"النموذج المحلي المضبوط حاليًا هو {self.model}"

    def _call_openai(self, prompt: str) -> str | None:
        if not self.openai_api_key:
            self._set_status("local", "OpenAI key missing")
            return None

        payload = {
            "model": self.openai_model,
            "input": prompt,
        }
        req = request.Request(
            self.openai_endpoint,
            data=json.dumps(payload).encode("utf-8"),
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {self.openai_api_key}",
            },
            method="POST",
        )

        try:
            with request.urlopen(req, timeout=45) as response:
                raw = response.read().decode("utf-8")
        except error.HTTPError as exc:
            try:
                error_body = exc.read().decode("utf-8")
            except Exception:
                error_body = ""
            details = error_body.strip() or exc.reason or exc.msg
            self._set_status("local", f"OpenAI HTTP {exc.code}: {details}")
            return None
        except (error.URLError, TimeoutError, OSError) as exc:
            self._set_status("local", f"OpenAI request failed: {exc.__class__.__name__}")
            return None

        try:
            envelope = json.loads(raw)
            output_text = envelope.get("output_text", "")
            if isinstance(output_text, str) and output_text.strip():
                self._set_status("openai")
                return output_text.strip()

            parts: list[str] = []
            for item in envelope.get("output", []):
                for content in item.get("content", []):
                    text_value = content.get("text")
                    if text_value:
                        parts.append(text_value)
            joined = "\n".join(parts).strip()
            if joined:
                self._set_status("openai")
            return joined or None
        except Exception:
            self._set_status("local", "OpenAI returned unreadable JSON")
            return None

    def answer_general(self, text: str, context: str = "") -> str | None:
        if self._should_use_online():
            answer = self._call_openai(self._build_general_prompt(text, context))
            if answer:
                return answer
            if self.mode == "online":
                self._set_status("local", self.last_error or "OpenAI request failed")

        payload = {
            "model": self.model,
            "prompt": self._build_general_prompt(text, context),
            "stream": False,
            "options": {"temperature": 0.5},
        }
        req = request.Request(
            self.endpoint,
            data=json.dumps(payload).encode("utf-8"),
            headers={"Content-Type": "application/json"},
            method="POST",
        )

        try:
            with request.urlopen(req, timeout=40) as response:
                raw = response.read().decode("utf-8")
        except (error.URLError, TimeoutError, OSError) as exc:
            self._set_status("local", f"Ollama request failed: {exc.__class__.__name__}")
            return None

        try:
            envelope = json.loads(raw)
            answer = (envelope.get("response", "") or "").strip()
            if answer:
                self._set_status("ollama")
            return answer or None
        except Exception:
            self._set_status("local", "Ollama returned unreadable JSON")
            return None

    def _build_prompt(self, text: str) -> str:
        today = date.today().isoformat()
        return f"""
أنت محلل أوامر عربي لتطبيق إداري اسمه جو.
تاريخ اليوم هو {today}.
أعد JSON فقط بدون أي شرح.

الصيغة:
{{
  "intent": "add_debt|add_bill|add_reminder|add_shopping_item|today_summary|unknown",
  "person_name": "",
  "vendor_name": "",
  "title": "",
  "item_name": "",
  "amount": 0,
  "currency": "USD",
  "due_date": "",
  "bill_date": "",
  "category": "",
  "notes": ""
}}

قواعد:
- إذا كان الطلب عن دين فـ intent = add_debt
- إذا كان الطلب عن فاتورة فـ intent = add_bill
- إذا كان الطلب عن تذكير فـ intent = add_reminder
- إذا كان الطلب عن مشتريات فـ intent = add_shopping_item
- إذا كان الطلب عن ملخص أو ماذا عندي اليوم فـ intent = today_summary
- إن لم تفهم فضع intent = unknown
- استخدم YYYY-MM-DD للتواريخ إن أمكن
- العملة USD للدولار وSYP لليرة

النص:
{text}
""".strip()

    def _build_general_prompt(self, text: str, context: str) -> str:
        today = date.today().isoformat()
        safe_context = context.strip() or "لا يوجد سياق إضافي."
        return f"""
أنت جو، مساعد ذكي محلي داخل تطبيق مكتبي إداري.
تاريخ اليوم هو {today}.
التطبيق يدير التذكيرات والديون والفواتير والمشتريات.
أولاً حدّد هل سؤال المستخدم عام أم متعلق ببيانات التطبيق.
إذا كان السؤال عاماً فأجب عليه كأي مساعد ذكي عادي ولا تتحدث عن بيانات التطبيق.
إذا كان السؤال عن بيانات التطبيق فاعتمد فقط على السياق التالي.
لا تقل إنك نفذت عملية داخل التطبيق إلا إذا كانت قد نُفذت فعلاً.
لا تبدأ بتحية عامة ولا تقل فقط "كيف أساعدك"، بل ادخل مباشرة في الإجابة.
اجعل الجواب مختصراً ما لم يطلب المستخدم التفصيل.

سياق التطبيق:
{safe_context}

رسالة المستخدم:
{text}
""".strip()


class JoeDesktopApp:
    def __init__(self) -> None:
        self.db = Database(DB_PATH)
        self.ai_parser = LocalAIParser()
        self.root = tk.Tk()
        self.root.title("Joe Desktop")
        self.root.geometry("1100x760")
        self.root.minsize(1000, 680)
        self.chat_messages: list[ChatMessage] = [
            ChatMessage("جو", "أهلًا سيدي. يمكنك أن تكتب لي أو تستخدم الميكروفون إن كان مدعومًا على هذا الجهاز.")
        ]
        self._build_ui()
        self.refresh_all()

    def run(self) -> None:
        self.root.mainloop()

    def _build_ui(self) -> None:
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        notebook = ttk.Notebook(self.root)
        notebook.grid(row=0, column=0, sticky="nsew")

        self.dashboard_tab = ttk.Frame(notebook, padding=14)
        self.reminders_tab = ttk.Frame(notebook, padding=14)
        self.debts_tab = ttk.Frame(notebook, padding=14)
        self.bills_tab = ttk.Frame(notebook, padding=14)
        self.shopping_tab = ttk.Frame(notebook, padding=14)
        self.chat_tab = ttk.Frame(notebook, padding=14)

        notebook.add(self.dashboard_tab, text="الملخص")
        notebook.add(self.reminders_tab, text="التذكيرات")
        notebook.add(self.debts_tab, text="الديون")
        notebook.add(self.bills_tab, text="الفواتير")
        notebook.add(self.shopping_tab, text="المشتريات")
        notebook.add(self.chat_tab, text="جو الصوتي")

        self._build_dashboard_tab()
        self._build_reminders_tab()
        self._build_debts_tab()
        self._build_bills_tab()
        self._build_shopping_tab()
        self._build_chat_tab()

    def _build_dashboard_tab(self) -> None:
        self.dashboard_tab.columnconfigure(0, weight=1)
        self.dashboard_tab.rowconfigure(1, weight=1)

        title = ttk.Label(
            self.dashboard_tab,
            text="جو - النسخة المكتبية",
            font=("Segoe UI", 20, "bold"),
        )
        title.grid(row=0, column=0, sticky="w", pady=(0, 12))

        self.dashboard_text = tk.Text(self.dashboard_tab, height=18, wrap="word", font=("Segoe UI", 12))
        self.dashboard_text.grid(row=1, column=0, sticky="nsew")

        actions = ttk.Frame(self.dashboard_tab)
        actions.grid(row=2, column=0, sticky="e", pady=(12, 0))

        ttk.Button(actions, text="نسخة احتياطية", command=self.export_backup).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(actions, text="تحديث الملخص", command=self.refresh_dashboard).grid(row=0, column=1)

    def _build_reminders_tab(self) -> None:
        frame = self.reminders_tab
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(4, weight=1)

        ttk.Label(frame, text="العنوان").grid(row=0, column=0, sticky="w", pady=4)
        self.reminder_title = ttk.Entry(frame)
        self.reminder_title.grid(row=0, column=1, sticky="ew", pady=4)

        ttk.Label(frame, text="التاريخ YYYY-MM-DD").grid(row=1, column=0, sticky="w", pady=4)
        self.reminder_due = ttk.Entry(frame)
        self.reminder_due.insert(0, date.today().isoformat())
        self.reminder_due.grid(row=1, column=1, sticky="ew", pady=4)

        ttk.Label(frame, text="ملاحظات").grid(row=2, column=0, sticky="w", pady=4)
        self.reminder_notes = ttk.Entry(frame)
        self.reminder_notes.grid(row=2, column=1, sticky="ew", pady=4)

        actions = ttk.Frame(frame)
        actions.grid(row=3, column=1, sticky="e", pady=(8, 10))

        ttk.Button(actions, text="حذف المحدد", command=self.delete_selected_reminder).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(actions, text="إضافة تذكير", command=self.add_reminder).grid(row=0, column=1)

        self.reminders_tree = ttk.Treeview(frame, columns=("title", "due", "notes"), show="headings")
        for col, title in (("title", "العنوان"), ("due", "التاريخ"), ("notes", "ملاحظات")):
            self.reminders_tree.heading(col, text=title)
            self.reminders_tree.column(col, width=220)
        self.reminders_tree.grid(row=4, column=0, columnspan=2, sticky="nsew")

    def _build_debts_tab(self) -> None:
        frame = self.debts_tab
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(6, weight=1)

        ttk.Label(frame, text="اسم الشخص").grid(row=0, column=0, sticky="w", pady=4)
        self.debt_person = ttk.Entry(frame)
        self.debt_person.grid(row=0, column=1, sticky="ew", pady=4)

        ttk.Label(frame, text="المبلغ").grid(row=1, column=0, sticky="w", pady=4)
        self.debt_amount = ttk.Entry(frame)
        self.debt_amount.grid(row=1, column=1, sticky="ew", pady=4)

        ttk.Label(frame, text="العملة").grid(row=2, column=0, sticky="w", pady=4)
        self.debt_currency = ttk.Entry(frame)
        self.debt_currency.insert(0, "USD")
        self.debt_currency.grid(row=2, column=1, sticky="ew", pady=4)

        ttk.Label(frame, text="الاستحقاق YYYY-MM-DD").grid(row=3, column=0, sticky="w", pady=4)
        self.debt_due = ttk.Entry(frame)
        self.debt_due.insert(0, date.today().isoformat())
        self.debt_due.grid(row=3, column=1, sticky="ew", pady=4)

        ttk.Label(frame, text="ملاحظات").grid(row=4, column=0, sticky="w", pady=4)
        self.debt_notes = ttk.Entry(frame)
        self.debt_notes.grid(row=4, column=1, sticky="ew", pady=4)

        actions = ttk.Frame(frame)
        actions.grid(row=5, column=1, sticky="e", pady=(8, 10))

        ttk.Button(actions, text="تحديد كمدفوع", command=self.mark_selected_debt_paid).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(actions, text="حذف المحدد", command=self.delete_selected_debt).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(actions, text="إضافة دين", command=self.add_debt).grid(row=0, column=2)

        self.debts_tree = ttk.Treeview(
            frame,
            columns=("person", "amount", "currency", "due", "status", "notes"),
            show="headings",
        )
        for col, title, width in (
            ("person", "الاسم", 180),
            ("amount", "المبلغ", 90),
            ("currency", "العملة", 80),
            ("due", "الاستحقاق", 110),
            ("status", "الحالة", 110),
            ("notes", "ملاحظات", 260),
        ):
            self.debts_tree.heading(col, text=title)
            self.debts_tree.column(col, width=width)
        self.debts_tree.grid(row=6, column=0, columnspan=2, sticky="nsew")

    def _build_bills_tab(self) -> None:
        frame = self.bills_tab
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(6, weight=1)

        ttk.Label(frame, text="اسم البائع").grid(row=0, column=0, sticky="w", pady=4)
        self.bill_vendor = ttk.Entry(frame)
        self.bill_vendor.grid(row=0, column=1, sticky="ew", pady=4)

        ttk.Label(frame, text="المبلغ").grid(row=1, column=0, sticky="w", pady=4)
        self.bill_amount = ttk.Entry(frame)
        self.bill_amount.grid(row=1, column=1, sticky="ew", pady=4)

        ttk.Label(frame, text="العملة").grid(row=2, column=0, sticky="w", pady=4)
        self.bill_currency = ttk.Entry(frame)
        self.bill_currency.insert(0, "USD")
        self.bill_currency.grid(row=2, column=1, sticky="ew", pady=4)

        ttk.Label(frame, text="التاريخ YYYY-MM-DD").grid(row=3, column=0, sticky="w", pady=4)
        self.bill_date = ttk.Entry(frame)
        self.bill_date.insert(0, date.today().isoformat())
        self.bill_date.grid(row=3, column=1, sticky="ew", pady=4)

        ttk.Label(frame, text="الفئة").grid(row=4, column=0, sticky="w", pady=4)
        self.bill_category = ttk.Entry(frame)
        self.bill_category.insert(0, "أسمدة زراعية")
        self.bill_category.grid(row=4, column=1, sticky="ew", pady=4)

        actions = ttk.Frame(frame)
        actions.grid(row=5, column=1, sticky="e", pady=(8, 10))

        ttk.Button(actions, text="تصدير Excel", command=self.export_bills_to_excel).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(actions, text="حذف المحدد", command=self.delete_selected_bill).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(actions, text="إضافة فاتورة", command=self.add_bill).grid(row=0, column=2)

        self.bills_tree = ttk.Treeview(
            frame,
            columns=("vendor", "amount", "currency", "date", "category"),
            show="headings",
        )
        for col, title, width in (
            ("vendor", "البائع", 220),
            ("amount", "المبلغ", 90),
            ("currency", "العملة", 80),
            ("date", "التاريخ", 110),
            ("category", "الفئة", 200),
        ):
            self.bills_tree.heading(col, text=title)
            self.bills_tree.column(col, width=width)
        self.bills_tree.grid(row=6, column=0, columnspan=2, sticky="nsew")

    def _build_shopping_tab(self) -> None:
        frame = self.shopping_tab
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(3, weight=1)

        ttk.Label(frame, text="العنصر").grid(row=0, column=0, sticky="w", pady=4)
        self.shopping_item = ttk.Entry(frame)
        self.shopping_item.grid(row=0, column=1, sticky="ew", pady=4)

        ttk.Label(frame, text="أضيف بواسطة").grid(row=1, column=0, sticky="w", pady=4)
        self.shopping_added_by = ttk.Entry(frame)
        self.shopping_added_by.insert(0, "البيت")
        self.shopping_added_by.grid(row=1, column=1, sticky="ew", pady=4)

        actions = ttk.Frame(frame)
        actions.grid(row=2, column=1, sticky="e", pady=(8, 10))

        ttk.Button(actions, text="تم الشراء", command=self.mark_selected_shopping_done).grid(row=0, column=0, padx=(0, 8))
        ttk.Button(actions, text="حذف المحدد", command=self.delete_selected_shopping_item).grid(row=0, column=1, padx=(0, 8))
        ttk.Button(actions, text="إضافة عنصر", command=self.add_shopping_item).grid(row=0, column=2)

        self.shopping_tree = ttk.Treeview(frame, columns=("item", "added_by"), show="headings")
        for col, title, width in (
            ("item", "العنصر", 260),
            ("added_by", "أضيف بواسطة", 160),
        ):
            self.shopping_tree.heading(col, text=title)
            self.shopping_tree.column(col, width=width)
        self.shopping_tree.grid(row=3, column=0, columnspan=2, sticky="nsew")

    def _build_chat_tab(self) -> None:
        frame = self.chat_tab
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(2, weight=1)

        ttk.Label(frame, text="جو - محادثة نصية وصوتية", font=("Segoe UI", 18, "bold")).grid(
            row=0, column=0, sticky="w", pady=(0, 10)
        )

        input_row = ttk.Frame(frame)
        input_row.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        input_row.columnconfigure(0, weight=1)

        self.chat_input = ttk.Entry(input_row)
        self.chat_input.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        self.ai_mode_var = tk.StringVar(value=getattr(self.ai_parser, "mode", "auto"))
        self.ai_mode_combo = ttk.Combobox(
            input_row,
            textvariable=self.ai_mode_var,
            values=("auto", "online", "local"),
            state="readonly",
            width=10,
        )
        self.ai_mode_combo.grid(row=0, column=1, padx=(0, 8))
        self.ai_mode_combo.bind("<<ComboboxSelected>>", self._on_ai_mode_changed)

        ttk.Button(input_row, text="إرسال", command=self.send_chat_text).grid(row=0, column=2, padx=(0, 8))
        ttk.Button(input_row, text="ميكروفون", command=self.listen_from_microphone).grid(row=0, column=3, padx=(0, 8))
        ttk.Button(input_row, text="أعد نطق آخر رد", command=self.repeat_last_joe_message).grid(row=0, column=4)

        self.ai_status_var = tk.StringVar(value=self._get_ai_status_text())
        ttk.Label(frame, textvariable=self.ai_status_var, foreground="#0b6b3a").grid(
            row=1, column=0, sticky="e", pady=(36, 0)
        )

        self.chat_text = tk.Text(frame, height=20, wrap="word", font=("Segoe UI", 12))
        self.chat_text.grid(row=2, column=0, sticky="nsew")
        self.chat_text.configure(state="disabled")

    def _validate_date(self, value: str) -> str:
        try:
            return datetime.strptime(value.strip(), "%Y-%m-%d").date().isoformat()
        except ValueError as exc:
            raise ValueError("التاريخ يجب أن يكون بالشكل YYYY-MM-DD") from exc

    def _clear_tree(self, tree: ttk.Treeview) -> None:
        for item in tree.get_children():
            tree.delete(item)

    def _get_selected_id(self, tree: ttk.Treeview, empty_message: str) -> int:
        selected = tree.selection()
        if not selected:
            raise ValueError(empty_message)
        return int(selected[0])

    def add_reminder(self) -> None:
        try:
            title = self.reminder_title.get().strip()
            if not title:
                raise ValueError("عنوان التذكير مطلوب")
            due_date = self._validate_date(self.reminder_due.get())
            self.db.add_reminder(title, due_date, self.reminder_notes.get().strip())
            self.reminder_title.delete(0, tk.END)
            self.reminder_notes.delete(0, tk.END)
            self.refresh_all()
        except Exception as exc:
            messagebox.showerror("خطأ", str(exc))

    def add_debt(self) -> None:
        try:
            person_name = self.debt_person.get().strip()
            if not person_name:
                raise ValueError("اسم الشخص مطلوب")
            amount = float(self.debt_amount.get().strip())
            due_date = self._validate_date(self.debt_due.get())
            currency = self.debt_currency.get().strip() or "USD"
            self.db.add_debt(person_name, amount, currency, due_date, self.debt_notes.get().strip())
            self.debt_person.delete(0, tk.END)
            self.debt_amount.delete(0, tk.END)
            self.debt_notes.delete(0, tk.END)
            self.refresh_all()
        except Exception as exc:
            messagebox.showerror("خطأ", str(exc))

    def add_bill(self) -> None:
        try:
            vendor_name = self.bill_vendor.get().strip()
            if not vendor_name:
                raise ValueError("اسم البائع مطلوب")
            amount = float(self.bill_amount.get().strip())
            bill_date = self._validate_date(self.bill_date.get())
            currency = self.bill_currency.get().strip() or "USD"
            category = self.bill_category.get().strip() or "غير مصنف"
            self.db.add_bill(vendor_name, amount, currency, bill_date, category)
            self.bill_vendor.delete(0, tk.END)
            self.bill_amount.delete(0, tk.END)
            self.refresh_all()
        except Exception as exc:
            messagebox.showerror("خطأ", str(exc))

    def add_shopping_item(self) -> None:
        try:
            item_name = self.shopping_item.get().strip()
            if not item_name:
                raise ValueError("اسم العنصر مطلوب")
            self.db.add_shopping_item(item_name, self.shopping_added_by.get().strip())
            self.shopping_item.delete(0, tk.END)
            self.refresh_all()
        except Exception as exc:
            messagebox.showerror("خطأ", str(exc))

    def delete_selected_reminder(self) -> None:
        try:
            reminder_id = self._get_selected_id(self.reminders_tree, "اختر تذكيرًا أولًا")
            self.db.delete_reminder(reminder_id)
            self.refresh_all()
        except Exception as exc:
            messagebox.showerror("خطأ", str(exc))

    def mark_selected_debt_paid(self) -> None:
        try:
            debt_id = self._get_selected_id(self.debts_tree, "اختر دينًا أولًا")
            self.db.mark_debt_paid(debt_id)
            self.refresh_all()
        except Exception as exc:
            messagebox.showerror("خطأ", str(exc))

    def delete_selected_debt(self) -> None:
        try:
            debt_id = self._get_selected_id(self.debts_tree, "اختر دينًا أولًا")
            self.db.delete_debt(debt_id)
            self.refresh_all()
        except Exception as exc:
            messagebox.showerror("خطأ", str(exc))

    def delete_selected_bill(self) -> None:
        try:
            bill_id = self._get_selected_id(self.bills_tree, "اختر فاتورة أولًا")
            self.db.delete_bill(bill_id)
            self.refresh_all()
        except Exception as exc:
            messagebox.showerror("خطأ", str(exc))

    def mark_selected_shopping_done(self) -> None:
        try:
            item_id = self._get_selected_id(self.shopping_tree, "اختر عنصرًا أولًا")
            self.db.mark_shopping_item_done(item_id)
            self.refresh_all()
        except Exception as exc:
            messagebox.showerror("خطأ", str(exc))

    def delete_selected_shopping_item(self) -> None:
        try:
            item_id = self._get_selected_id(self.shopping_tree, "اختر عنصرًا أولًا")
            self.db.delete_shopping_item(item_id)
            self.refresh_all()
        except Exception as exc:
            messagebox.showerror("خطأ", str(exc))

    def export_backup(self) -> None:
        try:
            backup_name = f"joe_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            destination = filedialog.asksaveasfilename(
                title="حفظ نسخة احتياطية",
                defaultextension=".db",
                initialfile=backup_name,
                filetypes=[("Database Files", "*.db"), ("All Files", "*.*")],
            )
            if not destination:
                return
            self.db.conn.commit()
            shutil.copyfile(DB_PATH, destination)
            messagebox.showinfo("نجاح", "تم حفظ النسخة الاحتياطية بنجاح")
        except Exception as exc:
            messagebox.showerror("خطأ", str(exc))

    def export_bills_to_excel(self) -> None:
        workbook_class = _load_workbook_class()
        if workbook_class is None:
            messagebox.showerror("خطأ", "ميزة Excel تحتاج مكتبة openpyxl غير المتوفرة حاليًا.")
            return

        try:
            default_name = f"joe_bills_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            destination = filedialog.asksaveasfilename(
                title="تصدير الفواتير إلى Excel",
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel Workbook", "*.xlsx")]
            )
            if not destination:
                return

            workbook = workbook_class()
            sheet = workbook.active
            sheet.title = "Bills"
            sheet.append(["ID", "Vendor", "Amount", "Currency", "Bill Date", "Category", "Paid"])

            total_unpaid = 0.0
            for row in self.db.fetch_all_bills():
                sheet.append([
                    row["id"],
                    row["vendor_name"],
                    row["amount"],
                    row["currency"],
                    row["bill_date"],
                    row["category"],
                    "Yes" if row["is_paid"] else "No",
                ])
                if not row["is_paid"]:
                    total_unpaid += float(row["amount"])

            summary = workbook.create_sheet("Summary")
            summary.append(["Metric", "Value"])
            summary.append(["Total Bills", len(self.db.fetch_all_bills())])
            summary.append(["Open Bills Amount", total_unpaid])
            summary.append(["Exported At", datetime.now().isoformat(timespec="seconds")])

            workbook.save(destination)
            messagebox.showinfo("نجاح", "تم تصدير الفواتير إلى Excel بنجاح")
        except Exception as exc:
            messagebox.showerror("خطأ", str(exc))

    def send_chat_text(self) -> None:
        user_text = self.chat_input.get().strip()
        if not user_text:
            messagebox.showerror("خطأ", "اكتب رسالة أولًا")
            return

        self.chat_input.delete(0, tk.END)
        self._handle_chat_message(user_text)

    def listen_from_microphone(self) -> None:
        sr, np, sd = _load_voice_input_modules()
        if sr is None or sd is None or np is None:
            messagebox.showinfo(
                "الميكروفون غير متاح",
                "ميزة الاستماع تحتاج مكتبات الصوت اللازمة. الرد الصوتي موجود، لكن الاستماع من الميكروفون غير متاح حاليًا في هذه النسخة."
            )
            return

        def worker() -> None:
            try:
                recognizer = sr.Recognizer()
                sample_rate = 16_000
                duration_seconds = 6
                self.root.after(0, lambda: self._append_chat("جو", "أستمع إليك الآن..."))
                recording = sd.rec(
                    int(duration_seconds * sample_rate),
                    samplerate=sample_rate,
                    channels=1,
                    dtype="int16",
                )
                sd.wait()
                audio_bytes = recording.tobytes()
                audio = sr.AudioData(audio_bytes, sample_rate, 2)
                transcript = recognizer.recognize_google(audio, language="ar")
                self.root.after(0, lambda: self._handle_chat_message(transcript))
            except Exception as exc:
                self.root.after(0, lambda: messagebox.showerror("تعذر الاستماع", str(exc)))

        threading.Thread(target=worker, daemon=True).start()

    def repeat_last_joe_message(self) -> None:
        for message in reversed(self.chat_messages):
            if message.sender == "جو":
                self._speak_text(message.content)
                return
        messagebox.showinfo("جو", "لا يوجد رد سابق ليتم نطقه")

    def _handle_chat_message(self, user_text: str) -> None:
        result: CommandResult | None = None
        if not self._looks_like_admin_request(user_text):
            ai_answer = self.ai_parser.answer_general(user_text, self._build_ai_context())
            if ai_answer:
                result = CommandResult(ai_answer)

        if result is None:
            result = JoeResponder(self.db, self.ai_parser).reply_to(user_text)

        self._append_chat("علاء", user_text)
        self._append_chat("جو", result.reply)
        self._speak_text(result.reply)
        if hasattr(self, "ai_status_var"):
            self.ai_status_var.set(self._get_ai_status_text())
        if result.refresh_ui:
            self.refresh_all()

    def _build_ai_context(self) -> str:
        snapshot = self.db.fetch_home_snapshot()
        debt_summary = self.db.fetch_summary()
        return (
            f"today_reminders={snapshot['today_reminders']}\n"
            f"open_bills={snapshot['open_bills']}\n"
            f"shopping_items={snapshot['shopping_items']}\n"
            f"due_today={debt_summary.due_today}\n"
            f"overdue={debt_summary.overdue}\n"
            f"total_open_amount={debt_summary.total_open_amount:.2f}"
        )

    def _get_ai_status_text(self) -> str:
        mode = getattr(self.ai_parser, "mode", "auto")
        last_provider = getattr(self.ai_parser, "last_provider", "")
        last_error = getattr(self.ai_parser, "last_error", "")
        if mode == "local":
            if last_provider == "ollama":
                return f"AI: Ollama active ({self.ai_parser.model})"
            return f"AI: Ollama ({self.ai_parser.model})"
        if mode == "online":
            if last_provider == "openai":
                return f"AI: OpenAI active ({self.ai_parser.openai_model})"
            if last_error:
                return f"AI: Online failed -> Ollama ({self.ai_parser.model}) | {last_error}"
            if getattr(self.ai_parser, "openai_api_key", ""):
                return f"AI: OpenAI ready ({self.ai_parser.openai_model})"
            return "AI: Online unavailable | OpenAI key missing"
        if getattr(self.ai_parser, "openai_api_key", ""):
            if last_provider == "openai":
                return f"AI: Auto -> OpenAI active ({self.ai_parser.openai_model})"
            if last_provider == "ollama":
                return f"AI: Auto -> Ollama active ({self.ai_parser.model})"
            return f"AI: Auto -> OpenAI ({self.ai_parser.openai_model}) / Ollama ({self.ai_parser.model})"
        if last_provider == "ollama":
            return f"AI: Auto -> Ollama active ({self.ai_parser.model})"
        return f"AI: Auto -> Ollama ({self.ai_parser.model})"

    def _on_ai_mode_changed(self, _event=None) -> None:
        selected_mode = self.ai_mode_var.get().strip().lower()
        if selected_mode not in {"auto", "online", "local"}:
            selected_mode = "auto"
        self.ai_parser.mode = selected_mode
        if hasattr(self, "ai_status_var"):
            self.ai_status_var.set(self._get_ai_status_text())

    def _looks_like_admin_request(self, user_text: str) -> bool:
        normalized = JoeResponder(self.db, self.ai_parser)._normalize(user_text)
        admin_keywords = (
            "دين",
            "فاتورة",
            "تذكير",
            "ذكرني",
            "مشتريات",
            "مشترياتي",
            "اضف",
            "أضف",
            "سجل",
            "ملخص",
            "اليوم",
        )
        return any(keyword in user_text or keyword in normalized for keyword in admin_keywords)

    def _append_chat(self, sender: str, content: str) -> None:
        self.chat_messages.append(ChatMessage(sender, content))
        self.chat_text.configure(state="normal")
        self.chat_text.insert(tk.END, f"{sender}: {content}\n\n")
        self.chat_text.see(tk.END)
        self.chat_text.configure(state="disabled")

    def _speak_text(self, text: str) -> None:
        def worker() -> None:
            escaped = text.replace("'", "''")
            command = (
                "Add-Type -AssemblyName System.Speech; "
                "$speaker = New-Object System.Speech.Synthesis.SpeechSynthesizer; "
                f"$speaker.Speak('{escaped}')"
            )
            try:
                subprocess.run(
                    ["powershell", "-NoProfile", "-Command", command],
                    check=False,
                    capture_output=True,
                    text=True,
                )
            except Exception:
                pass

        threading.Thread(target=worker, daemon=True).start()

    def refresh_dashboard(self) -> None:
        debt_summary = self.db.fetch_summary()
        snapshot = self.db.fetch_home_snapshot()

        text = (
            "ملخص سريع:\n\n"
            f"- تذكيرات اليوم: {snapshot['today_reminders']}\n"
            f"- ديون مستحقة اليوم: {debt_summary.due_today}\n"
            f"- ديون متأخرة: {debt_summary.overdue}\n"
            f"- إجمالي الديون المفتوحة: {debt_summary.total_open_amount:.2f}\n"
            f"- فواتير غير مدفوعة: {snapshot['open_bills']}\n"
            f"- عناصر مشتريات منزلية: {snapshot['shopping_items']}\n"
        )
        self.dashboard_text.delete("1.0", tk.END)
        self.dashboard_text.insert("1.0", text)

    def refresh_reminders(self) -> None:
        self._clear_tree(self.reminders_tree)
        for row in self.db.fetch_reminders():
            self.reminders_tree.insert(
                "",
                tk.END,
                iid=str(row["id"]),
                values=(row["title"], row["due_date"], row["notes"] or ""),
            )

    def refresh_debts(self) -> None:
        self._clear_tree(self.debts_tree)
        today = date.today().isoformat()
        for row in self.db.fetch_debts():
            due_date = row["due_date"]
            if row["is_paid"]:
                status = "مدفوع"
            elif due_date < today:
                status = "متأخر"
            elif due_date == today:
                status = "مستحق اليوم"
            else:
                status = "لاحقًا"

            self.debts_tree.insert(
                "",
                tk.END,
                iid=str(row["id"]),
                values=(
                    row["person_name"],
                    f"{row['amount']:.2f}",
                    row["currency"],
                    due_date,
                    status,
                    row["notes"] or "",
                ),
            )

    def refresh_bills(self) -> None:
        self._clear_tree(self.bills_tree)
        for row in self.db.fetch_bills():
            self.bills_tree.insert(
                "",
                tk.END,
                iid=str(row["id"]),
                values=(
                    row["vendor_name"],
                    f"{row['amount']:.2f}",
                    row["currency"],
                    row["bill_date"],
                    row["category"],
                ),
            )

    def refresh_shopping(self) -> None:
        self._clear_tree(self.shopping_tree)
        for row in self.db.fetch_shopping_items():
            self.shopping_tree.insert(
                "",
                tk.END,
                iid=str(row["id"]),
                values=(row["item_name"], row["added_by"] or ""),
            )

    def refresh_all(self) -> None:
        self.refresh_dashboard()
        self.refresh_reminders()
        self.refresh_debts()
        self.refresh_bills()
        self.refresh_shopping()
        if hasattr(self, "ai_status_var"):
            self.ai_status_var.set(self._get_ai_status_text())
        self.refresh_chat()

    def refresh_chat(self) -> None:
        self.chat_text.configure(state="normal")
        self.chat_text.delete("1.0", tk.END)
        for message in self.chat_messages:
            self.chat_text.insert(tk.END, f"{message.sender}: {message.content}\n\n")
        self.chat_text.configure(state="disabled")
        self.chat_text.see(tk.END)


class JoeResponder:
    def __init__(self, db: Database, ai_parser: LocalAIParser) -> None:
        self.db = db
        self.ai_parser = ai_parser

    def reply_to(self, text: str) -> CommandResult:
        input_text = text.strip()

        ai_result = self._try_local_ai_command(input_text)
        if ai_result is not None:
            return ai_result

        command_result = self._try_structured_command(input_text)
        if command_result is not None:
            return command_result

        normalized = self._normalize(input_text)

        if "شو عندي اليوم" in input_text or "ماذا عندي اليوم" in input_text:
            snapshot = self.db.fetch_home_snapshot()
            debt_summary = self.db.fetch_summary()
            return CommandResult(
                f"اليوم لديك {snapshot['today_reminders']} تذكيرات، و{debt_summary.due_today} ديون مستحقة اليوم، و{debt_summary.overdue} ديون متأخرة، و{snapshot['shopping_items']} عناصر مشتريات."
            )

        if "ملخص اليوم" in input_text or "شو صار معي اليوم" in input_text:
            snapshot = self.db.fetch_home_snapshot()
            return CommandResult(
                f"ملخص اليوم الحالي: {snapshot['today_reminders']} تذكيرات، {snapshot['open_bills']} فواتير غير مدفوعة، و{snapshot['shopping_items']} عناصر مشتريات."
            )

        if "حاله الذكاء" in normalized or "حالة الذكاء" in input_text:
            return CommandResult(
                f"أنا الآن أستخدم فهمًا قاعديًا محليًا مع دعم جاهز لربط نموذج محلي عبر Ollama. {self.ai_parser.availability_hint()}."
            )

        if "جو" in input_text or "مرحبا" in input_text or "صباح الخير" in input_text:
            return CommandResult("نعم سيدي، أنا معك.")

        return CommandResult("سمعتك يا سيدي. أستطيع الآن فهم عدة أوامر مثل تسجيل دين أو فاتورة أو تذكير أو عنصر مشتريات، كما أستطيع تصدير الفواتير إلى Excel.")

    def _try_local_ai_command(self, text: str) -> CommandResult | None:
        parsed = self.ai_parser.parse(text)
        if parsed is None or parsed.intent == "unknown":
            return None

        if parsed.intent == "add_debt" and parsed.person_name and parsed.amount > 0:
            due_date = parsed.due_date or date.today().isoformat()
            self.db.add_debt(parsed.person_name, parsed.amount, parsed.currency, due_date, parsed.notes)
            return CommandResult(
                f"فهمت الأمر عبر الذكاء المحلي. تم تسجيل دين على {parsed.person_name} بقيمة {parsed.amount:.2f} {parsed.currency} واستحقاق {due_date}.",
                refresh_ui=True,
            )

        if parsed.intent == "add_bill" and parsed.vendor_name and parsed.amount > 0:
            bill_date = parsed.bill_date or date.today().isoformat()
            self.db.add_bill(parsed.vendor_name, parsed.amount, parsed.currency, bill_date, parsed.category or "غير مصنف")
            return CommandResult(
                f"فهمت الأمر عبر الذكاء المحلي. تم تسجيل فاتورة للبائع {parsed.vendor_name} بقيمة {parsed.amount:.2f} {parsed.currency}.",
                refresh_ui=True,
            )

        if parsed.intent == "add_reminder" and parsed.title:
            due_date = parsed.due_date or date.today().isoformat()
            self.db.add_reminder(parsed.title, due_date, parsed.notes)
            return CommandResult(
                f"فهمت الأمر عبر الذكاء المحلي. تم تسجيل التذكير {parsed.title} بتاريخ {due_date}.",
                refresh_ui=True,
            )

        if parsed.intent == "add_shopping_item" and parsed.item_name:
            self.db.add_shopping_item(parsed.item_name, "جو")
            return CommandResult(
                f"فهمت الأمر عبر الذكاء المحلي. تمت إضافة {parsed.item_name} إلى قائمة المشتريات.",
                refresh_ui=True,
            )

        if parsed.intent == "today_summary":
            snapshot = self.db.fetch_home_snapshot()
            debt_summary = self.db.fetch_summary()
            return CommandResult(
                f"بحسب بيانات اليوم: {snapshot['today_reminders']} تذكيرات، {debt_summary.due_today} ديون مستحقة، {debt_summary.overdue} ديون متأخرة، و{snapshot['shopping_items']} عناصر مشتريات."
            )

        return None

    def _try_structured_command(self, text: str) -> CommandResult | None:
        normalized = self._normalize(text)

        if ("سجل" in normalized or "اضف" in normalized) and "دين" in normalized:
            parsed = self._parse_debt(normalized)
            if parsed is None:
                return CommandResult("فهمت أنك تريد تسجيل دين، لكنني أحتاج الاسم والمبلغ بشكل أوضح. مثال: سجل دين على أبو رامي 300 دولار بعد شهر.")

            self.db.add_debt(
                parsed["person_name"],
                parsed["amount"],
                parsed["currency"],
                parsed["due_date"],
                parsed["notes"],
            )
            return CommandResult(
                f"تم تسجيل دين على {parsed['person_name']} بقيمة {parsed['amount']:.2f} {parsed['currency']} وتاريخ استحقاق {parsed['due_date']}.",
                refresh_ui=True,
            )

        if ("سجل" in normalized or "اضف" in normalized) and "فاتورة" in normalized:
            parsed = self._parse_bill(normalized)
            if parsed is None:
                return CommandResult("فهمت أنك تريد تسجيل فاتورة، لكنني أحتاج اسم البائع والمبلغ بشكل أوضح. مثال: سجل فاتورة على الكيميائيات السورية 750 دولار فئة أسمدة زراعية.")

            self.db.add_bill(
                parsed["vendor_name"],
                parsed["amount"],
                parsed["currency"],
                parsed["bill_date"],
                parsed["category"],
            )
            return CommandResult(
                f"تم تسجيل فاتورة للبائع {parsed['vendor_name']} بقيمة {parsed['amount']:.2f} {parsed['currency']} بتاريخ {parsed['bill_date']}.",
                refresh_ui=True,
            )

        if ("سجل" in normalized or "اضف" in normalized) and ("تذكير" in normalized or "ذكرني" in normalized):
            parsed = self._parse_reminder(normalized)
            if parsed is None:
                return CommandResult("فهمت أنك تريد تذكيرًا، لكنني أحتاج نص التذكير بشكل أوضح. مثال: سجل تذكير طبيب الأسنان غدا.")

            self.db.add_reminder(parsed["title"], parsed["due_date"], parsed["notes"])
            return CommandResult(
                f"تم تسجيل التذكير: {parsed['title']} بتاريخ {parsed['due_date']}.",
                refresh_ui=True,
            )

        if ("اضف" in normalized or "سجل" in normalized) and ("مشتريات" in normalized or "مشتريات البيت" in normalized or "للقائمة" in normalized):
            parsed = self._parse_shopping_item(normalized)
            if parsed is None:
                return CommandResult("فهمت أنك تريد إضافة عنصر مشتريات، لكنني أحتاج اسم العنصر. مثال: اضف سكر إلى المشتريات.")
            self.db.add_shopping_item(parsed, "جو")
            return CommandResult(f"تمت إضافة {parsed} إلى قائمة المشتريات.", refresh_ui=True)

        if "صدر" in normalized and "اكسل" in normalized and "فاتور" in normalized:
            return CommandResult("أمر التصدير مفهوم. استخدم زر تصدير Excel داخل تبويب الفواتير في هذه النسخة الحالية.")

        return None

    def _parse_debt(self, text: str) -> dict[str, str | float] | None:
        amount_match = re.search(r"(\d+(?:\.\d+)?)", text)
        if amount_match is None:
            return None

        person_match = re.search(r"(?:على|ل|لـ)\s+(.+?)\s+(?:ب|بـ)?\s*\d", text)
        if person_match is None:
            person_match = re.search(r"دين\s+(.+?)\s+(?:ب|بـ)?\s*\d", text)
        if person_match is None:
            return None

        amount = float(amount_match.group(1))
        currency = "USD" if "دولار" in text else "SYP" if "ليرة" in text else "USD"
        due_date = self._extract_relative_date(text)
        notes = self._extract_notes(text)

        return {
            "person_name": person_match.group(1).strip(" .،"),
            "amount": amount,
            "currency": currency,
            "due_date": due_date,
            "notes": notes,
        }

    def _parse_bill(self, text: str) -> dict[str, str | float] | None:
        amount_match = re.search(r"(\d+(?:\.\d+)?)", text)
        if amount_match is None:
            return None

        vendor_match = re.search(r"(?:على|من)\s+(.+?)\s+\d", text)
        if vendor_match is None:
            vendor_match = re.search(r"فاتورة\s+(.+?)\s+\d", text)
        if vendor_match is None:
            return None

        category_match = re.search(r"فئة\s+(.+)$", text)
        category = category_match.group(1).strip(" .،") if category_match else "غير مصنف"
        currency = "USD" if "دولار" in text else "SYP" if "ليرة" in text else "USD"

        return {
            "vendor_name": vendor_match.group(1).strip(" .،"),
            "amount": float(amount_match.group(1)),
            "currency": currency,
            "bill_date": self._extract_relative_date(text),
            "category": category,
        }

    def _parse_reminder(self, text: str) -> dict[str, str] | None:
        title = text
        for prefix in ("سجل تذكير", "اضف تذكير", "ذكرني", "سجل", "اضف"):
            title = title.replace(prefix, "").strip()
        if not title:
            return None
        return {
            "title": title.strip(" .،"),
            "due_date": self._extract_relative_date(text),
            "notes": "",
        }

    def _parse_shopping_item(self, text: str) -> str | None:
        match = re.search(r"(?:اضف|سجل)\s+(.+?)\s+(?:الى|إلى)?\s*(?:المشتريات|القائمة|مشتريات البيت)", text)
        if match:
            return match.group(1).strip(" .،")
        return None

    def _extract_notes(self, text: str) -> str:
        match = re.search(r"(?:ملاحظة|ملاحظ[هة])\s*[:：]?\s*(.+)$", text)
        return match.group(1).strip() if match else ""

    def _extract_relative_date(self, text: str) -> str:
        today = date.today()
        if "بعد شهر" in text or "بعد شهرين" in text:
            days = 30 if "بعد شهر" in text else 60
            return (today + timedelta(days=days)).isoformat()
        if "غدا" in text or "بكرا" in text:
            return (today + timedelta(days=1)).isoformat()
        if "اليوم" in text:
            return today.isoformat()

        explicit = re.search(r"(\d{4}-\d{2}-\d{2})", text)
        if explicit:
            return explicit.group(1)
        return today.isoformat()

    def _normalize(self, text: str) -> str:
        return (
            text.strip()
            .replace("أ", "ا")
            .replace("إ", "ا")
            .replace("آ", "ا")
            .replace("ى", "ي")
            .replace("ة", "ه")
        )
